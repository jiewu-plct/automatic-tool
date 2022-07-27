from openpyxl import Workbook
from openpyxl import load_workbook
import os

file1='arm64.txt'
file2='riscv64.txt'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
filename = os.path.join(BASE_DIR, 'missing packages list.xlsx')

with open(file1, 'r') as f:
    arm64_data = f.readlines()

arm64_packagelist = [line.split('/')[0] for line in arm64_data]
# print('arm64_packagelist', arm64_packagelist)
print ('arm64_packagelist length', len(arm64_packagelist))

with open(file2, 'r') as f:
    riscv64_data = f.readlines()

riscv64_packagelist = [line.split('/')[0] for line in riscv64_data]
# print('riscv64_packagelist', riscv64_packagelist)
print ('riscv64_packagelist length', len(riscv64_packagelist))

wb = Workbook()
ws = wb.active

ws.title = "missing packages in riscv64"
wb.create_sheet("missing packages in arm64",1)
wb.save(filename)
wb = load_workbook(filename)
print ('sheetnames', wb.sheetnames)
ws = wb.worksheets[0]
ws.column_dimensions["A"].width = 45
riscv64_miss_packagelist = list(x for x in arm64_packagelist if x not in riscv64_packagelist)
arm64_miss_packagelist = list(x for x in riscv64_packagelist if x not in arm64_packagelist)
print ('riscv64_miss_packagelist length', len(riscv64_miss_packagelist))
print ('arm64_miss_packagelist length', len(arm64_miss_packagelist))
for i in range(len(riscv64_miss_packagelist)):
    ws.cell(i+1,1).value = riscv64_miss_packagelist[i]
ws = wb.worksheets[1]
ws.column_dimensions["A"].width = 40
for j in range(len(arm64_miss_packagelist)):
    ws.cell(j+1,1).value = arm64_miss_packagelist[j]

wb.save(filename)


